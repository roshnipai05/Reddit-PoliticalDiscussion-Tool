"""Evaluate the RAG system against the ground-truth QA set.

Reads:
    scripts/eval_set.json                 (ground-truth QA pairs)
    data/chroma_db/                       (built by build_index.py)

Writes:
    outputs/eval_raw_answers.json         (all questions + model answers)
    outputs/eval_scores.xlsx              (ROUGE-L + BERTScore + manual flags; editable Excel)
    outputs/evaluation_report.md          (final report, re-generated with --report-only)

Usage (two-step process):
    # Step 1: run pipeline, compute automatic metrics, produce Excel
    python scripts/evaluate_rag.py

    # Step 2: open outputs/eval_scores.xlsx, fill in the 0/1 flag columns
    #         (groq_faithful, groq_relevant), save the file, then:
    python scripts/evaluate_rag.py --report-only

Packages required:
    pip install rouge-score bert-score pandas openpyxl

Notes:
    - Gemini is disabled (--skip-gemini is forced on). Re-enable by removing
      the forced override in main() once your quota is restored.
    - BERTScore uses distilbert-base-uncased for speed; swap model_type to
      "roberta-large" in compute_bertscore() for higher quality at the cost
      of ~3× more time.
    - ROUGE-L and BERTScore are skipped (score = 0.0) for empty answers so
      they don't crash; check the no_answer_flag column to understand why an
      answer is empty.
"""

from __future__ import annotations

import argparse
import asyncio
import json
import time
from pathlib import Path
from typing import Any

import pandas as pd


ROOT = Path(__file__).resolve().parents[1]
EVAL_SET_PATH    = ROOT / "scripts"  / "eval_set.json"
CHROMA_DIR       = ROOT / "data"     / "chroma_db"
MODEL_CACHE_DIR  = ROOT / "data"     / "models" / "huggingface"
OUTPUTS_DIR      = ROOT / "outputs"

RAW_ANSWERS_PATH = OUTPUTS_DIR / "eval_raw_answers.json"
EXCEL_PATH       = OUTPUTS_DIR / "eval_scores.xlsx"
REPORT_PATH      = OUTPUTS_DIR / "evaluation_report.md"

# Groq free tier: ~6 000 tokens/min.  A short pause between sequential calls
# prevents hitting the per-minute limit when only Groq is active.
GROQ_INTER_CALL_SLEEP = 2.0


# ---------------------------------------------------------------------------
# Argument parsing
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--report-only", action="store_true",
        help="Skip pipeline run; reload existing raw answers + Excel and "
             "regenerate the Markdown report.",
    )
    parser.add_argument("--eval-set",       default=str(EVAL_SET_PATH))
    parser.add_argument("--chroma-dir",     default=str(CHROMA_DIR))
    parser.add_argument("--model-cache-dir",default=str(MODEL_CACHE_DIR))
    parser.add_argument("--skip-groq",   action="store_true")
    parser.add_argument("--skip-gemini", action="store_true",
                        help="Skip Gemini calls (forced True until quota is restored).")
    return parser.parse_args()


# ---------------------------------------------------------------------------
# Metric computation
# ---------------------------------------------------------------------------

def compute_rouge_l(reference: str, generated: str) -> float:
    """ROUGE-L F1.  Returns 0.0 for empty generated strings."""
    if not generated.strip():
        return 0.0
    from rouge_score import rouge_scorer  # type: ignore
    scorer = rouge_scorer.RougeScorer(["rougeL"], use_stemmer=True)
    return round(float(scorer.score(reference, generated)["rougeL"].fmeasure), 4)


def compute_bertscore(references: list[str], generateds: list[str]) -> list[float]:
    """
    BERTScore F1 for a batch of pairs.

    Uses distilbert-base-uncased by default (fast, no GPU needed).
    Empty generated strings receive a score of 0.0 and are excluded from the
    actual bert_score call so the library never crashes on blank input.
    """
    from bert_score import score as _bert_score  # type: ignore

    scores: list[float] = []
    valid_refs, valid_gens, valid_idx = [], [], []

    for i, (ref, gen) in enumerate(zip(references, generateds)):
        if gen.strip():
            valid_refs.append(ref)
            valid_gens.append(gen)
            valid_idx.append(i)
        else:
            scores.append(0.0)          # placeholder; will be re-inserted

    if valid_gens:
        _, _, F1 = _bert_score(
            valid_gens,
            valid_refs,
            lang="en",
            model_type="distilbert-base-uncased",   # swap to roberta-large for quality
            verbose=False,
        )
        f1_list = [round(float(f), 4) for f in F1.tolist()]
    else:
        f1_list = []

    # Re-merge: fill placeholders back in order
    result = [0.0] * len(references)
    empty_iter = iter([i for i in range(len(references)) if i not in valid_idx])
    valid_iter = iter(zip(valid_idx, f1_list))
    for i in range(len(references)):
        if i in valid_idx:
            _, v = next(valid_iter)
            result[i] = v
        # else already 0.0

    return result


# ---------------------------------------------------------------------------
# Pipeline execution
# ---------------------------------------------------------------------------

def load_eval_set(path: Path) -> list[dict[str, Any]]:
    return json.loads(path.read_text(encoding="utf-8"))


async def run_all_questions(
    eval_set: list[dict[str, Any]],
    post_collection,
    comment_collection,
    model_cache_dir: str,
    skip_groq: bool,
    skip_gemini: bool,
) -> list[dict[str, Any]]:
    """Run every question through the RAG pipeline; return list of result dicts."""
    from rag_query import retrieve, format_context, build_prompt  # type: ignore
    from rag_query import query_groq                               # type: ignore
    # query_gemini intentionally not imported while Gemini is disabled.

    results = []
    total = len(eval_set)

    for i, item in enumerate(eval_set, start=1):
        qid       = item["id"]
        question  = item["question"]
        qtype     = item["type"]
        answerable= item["answerable"]

        print(f"[{i}/{total}] {qid} ({qtype}) — {question[:70]}...")

        retrieval = retrieve(question, post_collection, comment_collection, model_cache_dir)
        context   = format_context(retrieval["posts"], retrieval["comments_by_post"])
        system, user = build_prompt(question, context, retrieval["no_answer_flag"])

        groq_answer   = ""
        gemini_answer = ""   # disabled

        if not skip_groq:
            try:
                groq_answer = query_groq(system, user)
            except Exception as exc:
                groq_answer = f"[error: {exc}]"
            time.sleep(GROQ_INTER_CALL_SLEEP)

        # ── Gemini is disabled ────────────────────────────────────────────────
        # To re-enable once your quota is restored:
        #   1. Remove the `skip_gemini = True` override in main().
        #   2. Uncomment the block below.
        #
        # if not skip_gemini:
        #     from rag_query import query_gemini
        #     try:
        #         gemini_answer = query_gemini(system, user)
        #     except Exception as exc:
        #         gemini_answer = f"[error: {exc}]"
        # ─────────────────────────────────────────────────────────────────────

        results.append({
            "id":                qid,
            "type":              qtype,
            "answerable":        answerable,
            "question":          question,
            "reference_answer":  item["reference_answer"],
            "no_answer_flag":    retrieval["no_answer_flag"],
            "max_cosine_sim":    round(retrieval["max_cosine_sim"], 4),
            "retrieved_post_ids":[p["post_id"] for p in retrieval["posts"]],
            "groq_answer":       groq_answer,
            "gemini_answer":     gemini_answer,
        })

    return results


# ---------------------------------------------------------------------------
# Scoring → produces a flat list of row dicts
# ---------------------------------------------------------------------------

def score_results(results: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """
    Compute ROUGE-L and BERTScore for each (model, question) pair.
    Manual flag columns are left blank for human annotation.
    Returns a list of row dicts (one per question).
    """
    references     = [r["reference_answer"] for r in results]
    groq_answers   = [r["groq_answer"]       for r in results]
    gemini_answers = [r["gemini_answer"]     for r in results]

    print("Computing ROUGE-L scores...")
    groq_rouge   = [compute_rouge_l(ref, gen) for ref, gen in zip(references, groq_answers)]
    gemini_rouge = [compute_rouge_l(ref, gen) for ref, gen in zip(references, gemini_answers)]

    print("Computing BERTScore (this may take a minute on first run while the model downloads)...")
    groq_bert   = compute_bertscore(references, groq_answers)
    gemini_bert = compute_bertscore(references, gemini_answers)

    rows = []
    for r, gr_rouge, gm_rouge, gr_bert, gm_bert in zip(
        results, groq_rouge, gemini_rouge, groq_bert, gemini_bert
    ):
        rows.append({
            # ── identity ───────────────────────────────────────────────────
            "id":               r["id"],
            "type":             r["type"],
            "answerable":       r["answerable"],
            "question":         r["question"],
            # ── reference (editable in Excel) ─────────────────────────────
            "reference_answer": r["reference_answer"],
            # ── retrieval diagnostics ──────────────────────────────────────
            "no_answer_flag":   r["no_answer_flag"],
            "max_cosine_sim":   r["max_cosine_sim"],
            "retrieved_post_ids": ", ".join(r["retrieved_post_ids"]),
            # ── model outputs ─────────────────────────────────────────────
            "groq_answer":      r["groq_answer"],
            "gemini_answer":    r["gemini_answer"],
            # ── automatic metrics ─────────────────────────────────────────
            "groq_rouge_l":          gr_rouge,
            "gemini_rouge_l":        gm_rouge,
            "groq_bertscore_f1":     gr_bert,
            "gemini_bertscore_f1":   gm_bert,
            # ── manual annotation (blank → fill with 0 or 1) ──────────────
            # faithfulness: 1 = answer is grounded in retrieved context
            #               0 = hallucination or unsupported claim detected
            # relevant:     1 = answer addresses the question
            #               0 = off-topic, or refused when it shouldn't have
            # For adversarial questions a correct refusal scores 1 for both.
            "groq_faithful":    "",
            "gemini_faithful":  "",
            "groq_relevant":    "",
            "gemini_relevant":  "",
            # ── free-text notes ───────────────────────────────────────────
            "notes":            "",
        })
    return rows


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def write_excel(score_rows: list[dict[str, Any]], path: Path) -> None:
    """
    Write a well-formatted .xlsx workbook with:
      Sheet 1 – Scores & Annotation  (metrics + yellow cells for manual flags)
      Sheet 2 – Full Answers          (complete answer text for review)
      Sheet 3 – Instructions
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.worksheet.datavalidation import DataValidation

    # ── style helpers ────────────────────────────────────────────────────────
    def font(size=10, bold=False, color="000000"):
        return Font(name="Arial", size=size, bold=bold, color=color)

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def border():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    def wrap(h="left", v="top"):
        return Alignment(horizontal=h, vertical=v, wrap_text=True)

    def center(v="center"):
        return Alignment(horizontal="center", vertical=v, wrap_text=True)

    NAVY, BLUE, LTBLUE = "1F3864", "2E75B6", "D6E4F7"
    AMBER, GREEN       = "FFF2CC", "E2EFDA"
    PINK, GREY, WHITE  = "FCE4D6", "F2F2F2", "FFFFFF"
    YELLOW             = "FFFF00"
    DISABLED_GREY      = "808080"

    wb = Workbook()

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 1 – Scores & Annotation
    # ════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Scores & Annotation"
    ws.freeze_panes = "A3"

    # column spec: (header, width, key)
    COLS = [
        ("ID",                         8,  "id"),
        ("Type",                       10, "type"),
        ("Answerable",                 13, "answerable"),
        ("Question",                   44, "question"),
        ("Reference Answer\n(editable)",52, "reference_answer"),
        ("No-Answer\nFlag",            11, "no_answer_flag"),
        ("Max Cosine\nSim",            11, "max_cosine_sim"),
        ("Retrieved\nPost IDs",        28, "retrieved_post_ids"),
        ("Groq\nAnswer",               56, "groq_answer"),
        ("Groq\nROUGE-L",             11, "groq_rouge_l"),
        ("Groq\nBERTScore F1",        13, "groq_bertscore_f1"),
        ("Groq\nFaithful\n(0/1)",     13, "groq_faithful"),
        ("Groq\nRelevant\n(0/1)",     13, "groq_relevant"),
        ("Gemini\nAnswer",             56, "gemini_answer"),
        ("Gemini\nROUGE-L",           11, "gemini_rouge_l"),
        ("Gemini\nBERTScore F1",      13, "gemini_bertscore_f1"),
        ("Gemini\nFaithful\n(0/1)",   13, "gemini_faithful"),
        ("Gemini\nRelevant\n(0/1)",   13, "gemini_relevant"),
        ("Notes",                      34, "notes"),
    ]

    GROQ_COLS   = set(range(9, 14))    # 1-indexed columns 9-13
    GEMINI_COLS = set(range(14, 19))   # 1-indexed columns 14-18

    # ── row 1: section banners ────────────────────────────────────────────────
    def banner(ws, r, c1, c2, text, bg):
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        cell = ws.cell(r, c1)
        cell.value     = text
        cell.font      = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        cell.fill      = fill(bg)
        cell.alignment = center()

    banner(ws, 1, 1,  8,  "RAG Evaluation  —  Scores & Manual Annotation",         NAVY)
    banner(ws, 1, 9,  13, "Groq  (llama-3.3-70b-versatile)",                       BLUE)
    banner(ws, 1, 14, 18, "Gemini  (gemini-2.0-flash)  —  disabled: quota exhausted", DISABLED_GREY)
    banner(ws, 1, 19, 19, "Notes", NAVY)
    ws.row_dimensions[1].height = 22

    # ── row 2: column headers ─────────────────────────────────────────────────
    for ci, (hdr, w, _) in enumerate(COLS, 1):
        c = ws.cell(2, ci)
        c.value     = hdr
        c.font      = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        c.alignment = center()
        c.border    = border()
        if ci in GROQ_COLS:
            c.fill = fill(BLUE)
        elif ci in GEMINI_COLS:
            c.fill = fill(DISABLED_GREY)
        else:
            c.fill = fill(NAVY)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 36

    # ── data validation for manual 0/1 cells ──────────────────────────────────
    dv = DataValidation(type="list", formula1='"0,1"', allow_blank=True,
                        showDropDown=False)
    dv.error      = "Enter 0 or 1"
    dv.errorTitle = "Invalid value"
    ws.add_data_validation(dv)

    MANUAL_KEYS = {"groq_faithful", "gemini_faithful",
                   "groq_relevant", "gemini_relevant"}
    METRIC_KEYS = {"groq_rouge_l", "gemini_rouge_l",
                   "groq_bertscore_f1", "gemini_bertscore_f1",
                   "max_cosine_sim"}
    WRAP_KEYS   = {"question", "reference_answer", "groq_answer",
                   "gemini_answer", "retrieved_post_ids", "notes"}

    for ri, row in enumerate(score_rows, start=3):
        is_adv  = not row["answerable"]
        row_bg  = PINK if is_adv else (GREY if ri % 2 == 0 else WHITE)

        for ci, (_, _, key) in enumerate(COLS, 1):
            c   = ws.cell(ri, ci)
            val = row.get(key)
            c.font   = font(9)
            c.border = border()

            # background
            if key == "reference_answer":
                c.fill = fill(AMBER)
            elif key in MANUAL_KEYS:
                c.fill = fill(YELLOW)
            elif ci in GROQ_COLS:
                c.fill = fill("EEF4FB" if not is_adv else "F9E0D5")
            elif ci in GEMINI_COLS:
                c.fill = fill("F5F5F5" if not is_adv else "F9E0D5")
            else:
                c.fill = fill(row_bg)

            # value + alignment + format
            if val is None or val == "":
                c.value = ""
            elif key == "answerable":
                c.value = "Yes" if val else "No  (adversarial)"
            elif key == "no_answer_flag":
                c.value = "TRUE" if val else "FALSE"
            elif key in METRIC_KEYS:
                c.value          = float(val)
                c.number_format  = "0.0000"
                c.alignment      = center("top")
            elif key in MANUAL_KEYS:
                c.value          = ""
                c.number_format  = "0"
                c.alignment      = center("top")
                dv.add(c)
            elif key in WRAP_KEYS:
                c.value     = str(val)
                c.alignment = wrap()
            else:
                c.value     = str(val)
                c.alignment = Alignment(horizontal="left", vertical="top")

        ws.row_dimensions[ri].height = 90

    # ── colour-scale conditional formatting on metric columns ─────────────────
    n = len(score_rows)
    for ci in [10, 11, 15, 16]:   # ROUGE-L and BERTScore columns
        col = get_column_letter(ci)
        ws.conditional_formatting.add(
            f"{col}3:{col}{2+n}",
            ColorScaleRule(
                start_type="min",  start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max",    end_color="63BE7B",
            ),
        )

    # ── summary / mean row ────────────────────────────────────────────────────
    sr = 3 + n
    ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=8)
    c = ws.cell(sr, 1)
    c.value     = "MEAN  (answerable questions only)"
    c.font      = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    c.fill      = fill(NAVY)
    c.alignment = center()

    answerable_rows = [i + 3 for i, r in enumerate(score_rows) if r["answerable"]]
    if answerable_rows:
        for ci, key in [(10, "groq_rouge_l"), (11, "groq_bertscore_f1"),
                        (15, "gemini_rouge_l"), (16, "gemini_bertscore_f1")]:
            col   = get_column_letter(ci)
            cells = ",".join(f"{col}{r}" for r in answerable_rows)
            c = ws.cell(sr, ci)
            c.value         = f"=AVERAGE({cells})"
            c.font          = Font(name="Arial", size=10, bold=True)
            c.fill          = fill(LTBLUE)
            c.number_format = "0.0000"
            c.alignment     = center()
            c.border        = border()

        for ci, key in [(12, "groq_faithful"), (13, "groq_relevant"),
                        (17, "gemini_faithful"), (18, "gemini_relevant")]:
            col   = get_column_letter(ci)
            cells = ",".join(f"{col}{r}" for r in answerable_rows)
            c = ws.cell(sr, ci)
            c.value         = f'=IFERROR(AVERAGE({cells}),"")'
            c.font          = Font(name="Arial", size=10, bold=True)
            c.fill          = fill(GREEN)
            c.number_format = "0.0%"
            c.alignment     = center()
            c.border        = border()
    ws.row_dimensions[sr].height = 22

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 2 – Full Answers (read-only reference)
    # ════════════════════════════════════════════════════════════════════════
    wa = wb.create_sheet("Full Answers")
    wa.freeze_panes = "A3"

    ANS_COLS = [
        ("ID",                          8),
        ("Type",                       10),
        ("Answerable",                 16),
        ("Question",                   52),
        ("Reference Answer\n(editable)", 62),
        ("Groq Answer",                66),
        ("Gemini Answer",              66),
    ]

    wa.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    c = wa.cell(1, 1)
    c.value     = "RAG Evaluation  —  Full Answers"
    c.font      = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    c.fill      = fill(NAVY)
    c.alignment = center()
    wa.row_dimensions[1].height = 22

    for ci, (hdr, w) in enumerate(ANS_COLS, 1):
        c = wa.cell(2, ci)
        c.value     = hdr
        c.font      = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        c.fill      = fill("7F6000" if ci == 5 else NAVY)
        c.alignment = center()
        c.border    = border()
        wa.column_dimensions[get_column_letter(ci)].width = w
    wa.row_dimensions[2].height = 28

    for ri, row in enumerate(score_rows, start=3):
        is_adv = not row["answerable"]
        vals   = [
            row["id"],
            row["type"],
            "No  (adversarial)" if is_adv else "Yes",
            row["question"],
            row["reference_answer"],
            row["groq_answer"],
            row["gemini_answer"] or "—  (Gemini disabled)",
        ]
        for ci, v in enumerate(vals, 1):
            c = wa.cell(ri, ci)
            c.value     = v
            c.font      = font(9)
            c.border    = border()
            c.alignment = wrap()
            if ci == 5:
                c.fill = fill(AMBER)
            elif is_adv:
                c.fill = fill(PINK)
            elif ri % 2 == 0:
                c.fill = fill(GREY)
            else:
                c.fill = fill(WHITE)
        wa.row_dimensions[ri].height = 110

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 3 – Instructions
    # ════════════════════════════════════════════════════════════════════════
    wi = wb.create_sheet("Instructions")
    wi.column_dimensions["A"].width = 90

    instructions = [
        ("RAG Evaluation Workbook  —  Instructions",         True,  14, NAVY),
        ("",                                                  False, 10, None),
        ("OVERVIEW",                                          True,  11, BLUE),
        ("This workbook has two data sheets:",               False, 10, None),
        ("  •  Scores & Annotation  – automatic metrics (ROUGE-L, BERTScore) + blank columns for your manual 0/1 flags", False, 10, None),
        ("  •  Full Answers  – complete model answers for reading and review", False, 10, None),
        ("",                                                  False, 10, None),
        ("STEP 1  —  Review answers",                         True,  10, None),
        ("  Open 'Full Answers' and read each Groq answer against the Reference Answer.", False, 10, None),
        ("",                                                  False, 10, None),
        ("STEP 2  —  Edit reference answers (optional)",      True,  10, None),
        ("  The amber Reference Answer column is editable. Update any ground-truth you want",  False, 10, None),
        ("  to improve, then re-run the script to recompute ROUGE-L / BERTScore.", False, 10, None),
        ("",                                                  False, 10, None),
        ("STEP 3  —  Fill in the yellow 0/1 flag cells in 'Scores & Annotation'", True, 10, None),
        ("  groq_faithful  / gemini_faithful",                False, 10, None),
        ("      1 = answer is grounded in the retrieved context  (no hallucination)", False, 10, None),
        ("      0 = hallucinated or unsupported claim detected", False, 10, None),
        ("  groq_relevant / gemini_relevant",                 False, 10, None),
        ("      1 = answer addresses the question",           False, 10, None),
        ("      0 = off-topic, or incorrectly refused",       False, 10, None),
        ("  For adversarial questions (pink rows) score 1 for both if the model",  False, 10, None),
        ("  correctly says the corpus does not contain the answer.", False, 10, None),
        ("",                                                  False, 10, None),
        ("STEP 4  —  Regenerate the Markdown report",         True,  10, None),
        ("  Save this workbook, then run:",                   False, 10, None),
        ("      python scripts/evaluate_rag.py --report-only", False, 10, None),
        ("  The report reads faithfulness % directly from this file.", False, 10, None),
        ("",                                                  False, 10, None),
        ("METRIC NOTES",                                      True,  11, BLUE),
        ("  ROUGE-L         Longest-common-subsequence F1. Range 0–1; higher = better lexical overlap.", False, 10, None),
        ("  BERTScore F1    Contextual embedding similarity (distilbert-base-uncased). Range 0–1.", False, 10, None),
        ("                  Swap to roberta-large in compute_bertscore() for higher quality.", False, 10, None),
        ("  Max Cosine Sim  Highest similarity between the query and any retrieved chunk.", False, 10, None),
        ("  No-Answer Flag  TRUE when max cosine sim < 0.35; the system declines to answer.", False, 10, None),
        ("",                                                  False, 10, None),
        ("COLOUR CODING",                                     True,  11, BLUE),
        ("  Amber           Reference Answer column — editable ground-truth", False, 10, None),
        ("  Yellow          Manual flag cells — fill these in",     False, 10, None),
        ("  Pink rows       Adversarial questions (expected refusals)", False, 10, None),
        ("  Green (summary) Faithfulness / relevance % (auto-computed once flags are filled)", False, 10, None),
        ("  Red→Yellow→Green colour scale on metric columns (low → mid → high)", False, 10, None),
        ("",                                                  False, 10, None),
        ("RE-ENABLING GEMINI",                                True,  11, BLUE),
        ("  Once your Gemini quota is restored:",             False, 10, None),
        ("  1.  Remove the line  skip_gemini = True  from main() in evaluate_rag.py", False, 10, None),
        ("  2.  Uncomment the Gemini block in run_all_questions()", False, 10, None),
        ("  3.  Re-run the script — Gemini answers + metrics will populate automatically", False, 10, None),
    ]

    for row_i, (text, bold, size, bg) in enumerate(instructions, start=1):
        c = wi.cell(row_i, 1)
        c.value     = text
        c.font      = Font(name="Arial", size=size, bold=bold,
                           color="FFFFFF" if bg else "000000")
        c.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True)
        if bg:
            c.fill = fill(bg)
        wi.row_dimensions[row_i].height = 18 if text else 8

    wb.save(str(path))
    print(f"Excel workbook saved: {path}")


# ---------------------------------------------------------------------------
# Report generation
# ---------------------------------------------------------------------------

def generate_report(
    score_rows: list[dict[str, Any]],
    raw_results: list[dict[str, Any]],
) -> str:
    """Build evaluation_report.md from score rows (list of dicts)."""

    scores_df = pd.DataFrame(score_rows)

    # Coerce numeric columns
    for col in ["groq_rouge_l", "gemini_rouge_l",
                "groq_bertscore_f1", "gemini_bertscore_f1"]:
        scores_df[col] = pd.to_numeric(scores_df[col], errors="coerce").fillna(0.0)

    has_manual = (
        scores_df["groq_faithful"].astype(str).str.strip().ne("").any()
    )

    lines = [
        "# RAG System Evaluation Report",
        "",
        "## 1. Configuration",
        "",
        "- **Embedding model**: sentence-transformers/all-mpnet-base-v2",
        "- **Vector store**: ChromaDB (collections: reddit_posts, reddit_comments)",
        "- **Retrieval**: top-5 posts (diversity-filtered) + up to 5 comments per post",
        "- **Re-ranking**: cosine similarity × log(1 + reddit_score)",
        "- **No-answer threshold**: cosine similarity < 0.35",
        "- **LLM A (Groq)**: llama-3.3-70b-versatile",
        "- **LLM B (Gemini)**: gemini-2.0-flash  _(disabled — quota exhausted)_",
        f"- **Evaluation set**: {len(scores_df)} questions "
        f"({scores_df['type'].value_counts().to_dict()})",
        "",
        "---",
        "",
        "## 2. Results Table",
        "",
    ]

    if has_manual:
        header = (
            "| ID | Type | Groq ROUGE-L | Gemini ROUGE-L | "
            "Groq BERTScore | Gemini BERTScore | "
            "Groq Faithful | Gemini Faithful | "
            "Groq Relevant | Gemini Relevant |"
        )
        sep = "|" + "|".join(["-" * 16] * 10) + "|"
    else:
        header = (
            "| ID | Type | Groq ROUGE-L | Gemini ROUGE-L | "
            "Groq BERTScore | Gemini BERTScore |"
        )
        sep = "|" + "|".join(["-" * 16] * 6) + "|"

    lines += [header, sep]

    for _, row in scores_df.iterrows():
        if has_manual:
            line = (
                f"| {row['id']} | {row['type']} "
                f"| {row['groq_rouge_l']:.4f} | {row['gemini_rouge_l']:.4f} "
                f"| {row['groq_bertscore_f1']:.4f} | {row['gemini_bertscore_f1']:.4f} "
                f"| {row['groq_faithful']} | {row['gemini_faithful']} "
                f"| {row['groq_relevant']} | {row['gemini_relevant']} |"
            )
        else:
            line = (
                f"| {row['id']} | {row['type']} "
                f"| {row['groq_rouge_l']:.4f} | {row['gemini_rouge_l']:.4f} "
                f"| {row['groq_bertscore_f1']:.4f} | {row['gemini_bertscore_f1']:.4f} |"
            )
        lines.append(line)

    # Summary row
    mean_cols = ["groq_rouge_l", "gemini_rouge_l",
                 "groq_bertscore_f1", "gemini_bertscore_f1"]
    means = scores_df[mean_cols].mean()

    if has_manual:
        def pct(col: str) -> str:
            vals = pd.to_numeric(scores_df[col], errors="coerce")
            return f"{vals.mean() * 100:.1f}%" if vals.notna().any() else "N/A"
        mean_row = (
            f"| **Mean** | — "
            f"| **{means['groq_rouge_l']:.4f}** | **{means['gemini_rouge_l']:.4f}** "
            f"| **{means['groq_bertscore_f1']:.4f}** | **{means['gemini_bertscore_f1']:.4f}** "
            f"| **{pct('groq_faithful')}** | **{pct('gemini_faithful')}** "
            f"| **{pct('groq_relevant')}** | **{pct('gemini_relevant')}** |"
        )
    else:
        mean_row = (
            f"| **Mean** | — "
            f"| **{means['groq_rouge_l']:.4f}** | **{means['gemini_rouge_l']:.4f}** "
            f"| **{means['groq_bertscore_f1']:.4f}** | **{means['gemini_bertscore_f1']:.4f}** |"
        )
    lines += [mean_row, ""]

    # Adversarial sub-table
    adversarial = scores_df[~scores_df["answerable"].astype(bool)]
    if not adversarial.empty:
        lines += [
            "---", "",
            "## 3. Adversarial Question Behaviour",
            "",
            "| ID | Question (truncated) | no_answer_flag | "
            "Groq Refused Correctly | Gemini Refused Correctly |",
            "|---|---|---|---|---|",
        ]
        raw_lookup = {r["id"]: r for r in raw_results}
        for _, row in adversarial.iterrows():
            raw           = raw_lookup.get(row["id"], {})
            q_short       = str(row["question"])[:60]
            flag          = "✓" if row["no_answer_flag"] else "✗"
            refusal_phrase= "does not contain sufficient information"
            groq_ans      = raw.get("groq_answer", "")
            gemini_ans    = raw.get("gemini_answer", "")
            groq_refused  = "✓" if refusal_phrase in groq_ans.lower()   else "✗ (check manually)"
            gemini_refused= "✓" if refusal_phrase in gemini_ans.lower() else "✗ (check manually)"
            lines.append(
                f"| {row['id']} | {q_short}... | {flag} "
                f"| {groq_refused} | {gemini_refused} |"
            )
        lines.append("")

    # Per-type breakdown
    lines += [
        "---", "",
        "## 4. Performance by Question Type",
        "",
        "| Type | Count | Groq ROUGE-L | Gemini ROUGE-L | "
        "Groq BERTScore | Gemini BERTScore |",
        "|---|---|---|---|---|---|",
    ]
    for qtype, group in scores_df.groupby("type"):
        lines.append(
            f"| {qtype} | {len(group)} "
            f"| {group['groq_rouge_l'].mean():.4f} "
            f"| {group['gemini_rouge_l'].mean():.4f} "
            f"| {group['groq_bertscore_f1'].mean():.4f} "
            f"| {group['gemini_bertscore_f1'].mean():.4f} |"
        )
    lines.append("")

    # Full answers
    lines += [
        "---", "",
        "## 5. Full Model Answers",
        "",
        "_Generated answers for manual review._",
        "",
    ]
    for r in raw_results:
        lines += [
            f"### {r['id']} — {r['type']} "
            f"{'(adversarial)' if not r['answerable'] else ''}",
            f"**Question**: {r['question']}",
            f"**Reference**: {r['reference_answer']}",
            f"**no_answer_flag**: {r['no_answer_flag']}  |  "
            f"**max_cosine_sim**: {r['max_cosine_sim']}",
            "",
            "**Groq answer**:",
            f"> {r['groq_answer']}",
            "",
            "**Gemini answer**:",
            f"> {r['gemini_answer'] or '_disabled_'}",
            "",
            "---",
            "",
        ]

    # Qualitative analysis placeholders
    lines += [
        "## 6. Qualitative Analysis",
        "",
        "_(Fill in after reviewing the full answers above.)_",
        "",
        "### Where Groq succeeds", "", "- ", "",
        "### Where Groq fails",    "", "- ", "",
        "### Where Gemini succeeds (once re-enabled)", "", "- ", "",
        "### Where Gemini fails",   "", "- ", "",
        "### Retrieval quality observations", "", "- ", "",
        "### Adversarial behaviour", "", "- ", "",
    ]

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    args = parse_args()

    # ── Gemini is disabled until quota is restored ────────────────────────────
    args.skip_gemini = True
    # ─────────────────────────────────────────────────────────────────────────

    OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
    eval_set = load_eval_set(Path(args.eval_set))

    # ── Report-only mode ──────────────────────────────────────────────────────
    if args.report_only:
        if not RAW_ANSWERS_PATH.exists():
            raise FileNotFoundError(
                f"Raw answers not found at {RAW_ANSWERS_PATH}. "
                "Run without --report-only first."
            )
        raw_results = json.loads(RAW_ANSWERS_PATH.read_text(encoding="utf-8"))

        if EXCEL_PATH.exists():
            # Read back the (possibly manually annotated) Excel
            df = pd.read_excel(EXCEL_PATH, sheet_name="Scores & Annotation",
                               header=1)   # row index 1 = our column-header row
            score_rows = df.to_dict(orient="records")
            print(f"Loaded scores from {EXCEL_PATH}")
        else:
            raise FileNotFoundError(
                f"Excel file not found at {EXCEL_PATH}. "
                "Run without --report-only first."
            )

        report = generate_report(score_rows, raw_results)
        REPORT_PATH.write_text(report, encoding="utf-8")
        print(f"Report written: {REPORT_PATH}")
        return

    # ── Full pipeline run ─────────────────────────────────────────────────────
    import sys
    sys.path.insert(0, str(ROOT / "scripts"))
    from rag_query import get_collections  # type: ignore  # noqa: E402

    post_collection, comment_collection = get_collections(args.chroma_dir)

    print(f"Running {len(eval_set)} questions through RAG pipeline...")
    raw_results = asyncio.run(
        run_all_questions(
            eval_set=eval_set,
            post_collection=post_collection,
            comment_collection=comment_collection,
            model_cache_dir=args.model_cache_dir,
            skip_groq=args.skip_groq,
            skip_gemini=args.skip_gemini,
        )
    )

    RAW_ANSWERS_PATH.write_text(
        json.dumps(raw_results, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    print(f"Raw answers saved: {RAW_ANSWERS_PATH}")

    print("Scoring...")
    score_rows = score_results(raw_results)

    write_excel(score_rows, EXCEL_PATH)

    report = generate_report(score_rows, raw_results)
    REPORT_PATH.write_text(report, encoding="utf-8")
    print(f"Preliminary report written: {REPORT_PATH}")

    print(
        "\nNext steps:\n"
        f"  1. Open {EXCEL_PATH}\n"
        "  2. Fill in the yellow groq_faithful / groq_relevant columns (0 or 1)\n"
        "     Pink rows = adversarial questions: score 1 if the model correctly refused\n"
        "  3. Save the file, then run:\n"
        "         python scripts/evaluate_rag.py --report-only\n"
        "  4. Update the reference_answer column (amber) if any ground-truth needs fixing,\n"
        "     then re-run the full pipeline to recompute ROUGE-L / BERTScore.\n"
    )


if __name__ == "__main__":
    main()